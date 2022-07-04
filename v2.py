from bitkub import Bitkub
from configparser import ConfigParser #การเรียกไฟล์ Config
import time
from openpyxl import load_workbook
file='log.xlsx'
lo=load_workbook(file)
log=lo.active
dbconf = ConfigParser()
dbconf.read_file(open(r'config.ini'))
KEY = dbconf.get('Config', 'API_KEY')
PASS = dbconf.get('Config', 'API_SECRET')
Line_Notify = dbconf.get('Config', 'LineNotify')
Asset = dbconf.get('Config', 'Asset').split(",")
coo = dbconf.get('Config', 'Core').split(",")
DCA = dbconf.get('Config', 'DCA')
GAP = float(dbconf.get('Config', 'GAP'))
CASH = float(dbconf.get('Config', 'Cash'))
timelo = int(dbconf.get('Config', 'Timeloop'))
log['A1']='Asset'
log['A2']='Core'
log['A3']='Befor'
Bot=" "

def cellbe(i):
    return log.cell(8,i+1).value
def cellas(i):
    return log.cell(6,i+1).value
def cellco(i):
    return log.cell(7,i+1).value
try:
    for i in range(len(Asset)):
        if log.cell(1, i + 2).value != Asset[i]:
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])
        elif log.cell(3, i + 2).value != int(coo[i]):
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])

    lo.save(file)
except Exception as e:
    print(e)
    print('Close in 60s')
    time.sleep(60)

from line_notify import LineNotify  # การส่ง Line
notify = LineNotify(Line_Notify)
bitkub = Bitkub(api_key=KEY,
                api_secret=PASS)
hr=0
while True :
    n = 0
    H = 60/timelo
    while (n < H):
        try:
            res = 'result'
            named_tuple = time.localtime()
            print('\n............................................\n')
            time_string = time.strftime("%d/%m/%Y %H:%M:%S", named_tuple)
            print(time_string)
            for i in range(len(Asset)):
                Get_balance = bitkub.wallet()
                Core = log.cell(2, i + 2).value
                Asset_01 = Get_balance[res][Asset[i]]
                AssetName = 'THB_' + Asset[i]
                get_price = bitkub.ticker(AssetName)
                Asset_01_Value = Asset_01 * get_price[AssetName]['last']
                print(Asset_01, Asset[i], '=', "{:.2f}".format(Asset_01_Value), '<==>', str(Core))
                rat = get_price[AssetName]['last']
                CoreAsset = int(Core)
                DiffAsset = (CoreAsset * GAP / 100)
                print('Diff = ' + '{:.2f}'.format(DiffAsset))

                if Asset_01_Value > (CoreAsset + DiffAsset):
                    diff_sell = CoreAsset * GAP / 100
                    bitkub.place_ask_by_fiat(sym=AssetName, amt=diff_sell, rat=rat, typ='market')
                    log[str(cellco(i + 1))] = int(Core + DCA)  # ขยายพอร์ตเมื่อมีการขาย ทีละ 2฿
                    lo.save(file)  # Save log
                    CoreSell = 'Sell ' + AssetName + ', Core = ฿' + str(Core + 2)
                    notify.send(CoreSell)
                    print("SELL " + str(diff_sell) + " ฿")

                elif Asset_01_Value < (CoreAsset - DiffAsset):
                    diff_buy = CoreAsset * GAP / 100
                    bitkub.place_bid(sym=AssetName, amt=diff_buy, rat=rat, typ='market')
                    print("Buy " + str(diff_buy) + " ฿")
                    notify.send('Buy ' + AssetName)

                else:
                    print('Diff balance = '"{:.2f}".format(Asset_01_Value - CoreAsset), '฿')
            balance = float(format(Get_balance[res]['THB']))
            P = balance - CASH
            #Bot = 'Cash ' + ('{:.2f}'.format(balance)+ " ฿")
            bot = '\n💰 Cash : ' + ('{:.2f}'.format(balance)+ " ฿" + '\n💸 Profit : ' + '{:.2f}'.format(P) + ' ฿')
            #print(Bot)
            #print("Profit {:.2f}".format(P)+ " ฿")
            print('\n............................................\n')
            time.sleep(60 *timelo)
        except Exception as e:
            print(e)
            try:
                notify.send(e)
            except Exception as e:
                print(e)
                pass
            pass
        n += 1
    notify.send(bot)
    hr+=1
    if(hr==24):
        hr=0
        for i in range(len(Asset)):
            Core = log.cell(2, i + 2).value + int(DCA)  # ขยายพอร์ตทุก 24 ชม. (DCA)
            print(Asset[i], str(Core))
            log[str(cellco(i + 1))] = Core
            lo.save(file)
            try:
                notify.send(Asset[i] + ' = ฿' + str(Core))
            except Exception as e:
                print(e)
                pass