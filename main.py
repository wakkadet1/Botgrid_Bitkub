from bitkub import Bitkub
from configparser import ConfigParser #การเรียกไฟล์ Config
import time
from colorama import Fore, Back, Style
from openpyxl import load_workbook
from colorama import *
from line_notify import LineNotify
file='log.xlsx'
lo=load_workbook(file)
log=lo.active
dbconf = ConfigParser()
dbconf.read_file(open(r'config.ini'))
KEY = dbconf.get('Config', 'API_KEY')
PASS = dbconf.get('Config', 'API_SECRET')
Line_Notify = dbconf.get('Config', 'LineNotify')
Asset = dbconf.get('Config', 'Asset').split(",")
coo = dbconf.get('Config', 'Core').split(",")
DCA = dbconf.get('Config', 'DCA')
GAPUP = float(dbconf.get('Config', 'GAPUP'))
GAPDOWN = float(dbconf.get('Config', 'GAPDOWN'))
CASH = float(dbconf.get('Config', 'Cash'))
timeli = int(dbconf.get('Config', 'Time'))
timelo = int(dbconf.get('Config', 'Timeloop'))
log['A1']='Asset'
log['A2']='Core'
log['A3']='Befor'
P=" "

def cellbe(i):
    return log.cell(8,i+1).value
def cellas(i):
    return log.cell(6,i+1).value
def cellco(i):
    return log.cell(7,i+1).value
try:
    for i in range(len(Asset)):
        if log.cell(1, i + 2).value != Asset[i]:
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])

        elif log.cell(3, i + 2).value != int(coo[i]):
            log[str(cellas(i + 1))] = Asset[i]
            log[str(cellco(i + 1))] = int(coo[i])
            log[str(cellbe(i + 1))] = int(coo[i])
    lo.save(file)
except Exception as e:
    print(e)
    print('Close in 60s')
    time.sleep(60)


init(autoreset=True)
# การส่ง Line
notify = LineNotify(Line_Notify)

try:
    bitkub = Bitkub()
    bitkub.set_api_key(KEY)
    bitkub.set_api_secret(PASS)
    bitkub.status()
    bitkub.servertime()
except Exception as e:
    try:
        notify.send(e)
    except Exception as e:
        pass
    pass

while True :
    h = 0
    while (h < 24):
        n = 0
        while (n < 60):
                # เริ่มการทำงานของบอท
            try:
                res = 'result'
                Get_balance = bitkub.wallet()
                print(Fore.RED + '\n##############################################')
                named_tuple = time.localtime()
                time_string = time.strftime("%d/%m/%Y %H:%M:%S", named_tuple)
                print(Fore.CYAN+time_string)
                for i in range(len(Asset)):

                    Core = log.cell(2, i + 2).value
                    Asset_01 = Get_balance[res][Asset[i]]
                    AssetName = 'THB_' + Asset[i]
                    get_price = bitkub.ticker(AssetName)
                    Asset_01_Value = Asset_01 * get_price[AssetName]['last']
                    print(Asset_01, Asset[i], '=', "{:.2f}".format(Asset_01_Value),' <==> ',Core)
                    rat = get_price[AssetName]['last']

                    CoreAsset = int(Core)

                    DiffDown = (CoreAsset * GAPDOWN/100)
                    DiffUp  = (CoreAsset * GAPUP/100)

                    print("DiffUp == ",'{:.2f}'.format(DiffUp))
                    print("DiffDown == ",'{:.2f}'.format(DiffDown))

                    if Asset_01_Value > (CoreAsset + DiffUp): # เงื่อนไขในการขาย
                        bitkub.place_ask_by_fiat(sym=AssetName, amt=DiffUp, rat=rat, typ='market')
                        log[str(cellco(i + 1))] = int(Core + DCA)  # ขยายพอร์ตเมื่อมีการขาย ทีละ
                        lo.save(file)  # Save log
                        CoreSell = 'Sell ' + AssetName + ', Sell = ' + '{:.2f}'.format(DiffUp)+' ฿'
                        notify.send(CoreSell)
                        print("SELL " + str(diff_sell) + " ฿")

                    elif Asset_01_Value < (CoreAsset - DiffDown): # เงื่อนไขในการซื้อ
                        bitkub.place_bid(sym=AssetName, amt=DiffDown, rat=rat, typ='market')
                        print("Buy " + str(DiffDown) + " ฿")
                        CoreBuy = 'Buy ' + AssetName + ', Buy = ' + '{:.2f}'.format(DiffDown) + ' ฿'
                        notify.send(CoreBuy)
                    else:
                        print(Fore.GREEN +Style.BRIGHT+ 'Diff '+"{:.2f}".format(Asset_01_Value - CoreAsset)+ ' ฿')
                n += timelo
                print(Fore.LIGHTBLUE_EX+'Cash',"{:.2f}".format(Get_balance[res]['THB']),' ฿')
                balance = float(format(Get_balance[res]['THB']))
                P = balance - CASH
                print(Fore.YELLOW + 'Profit','{:.2f}'.format(P),' ฿')
                print(Fore.RED + '\n##############################################')
                balancep = '\nCash '+ ('{:.2f}'.format(Get_balance[res]['THB'])) + ' ฿ '+ '\nProfit ' + '{:.2f}'.format(P) + ' ฿'
                time.sleep(timelo*60)

            except Exception as e:
                print(e)
                try:
                    notify.send(e)
                except Exception as e:
                    print(e)
                    pass
                pass
        h += 1
        a += 1
        if(a==timeli):
            notify.send(balancep)
            a=0

    for i in range(len(Asset)):
        Core = log.cell(2, i + 2).value + int(DCA)  # ขยายพอร์ตทุก 24 ชม. (DCA)
        print(Asset[i], str(Core))
        log[str(cellco(i + 1))]=Core
        lo.save(file)
        try:
            notify.send(Asset[i] + ' = ฿' + str(Core))
        except Exception as e:
            print(e)
            pass